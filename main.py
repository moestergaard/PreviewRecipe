import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.reader.excel import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
import tempfile
import subprocess
import requests

pdf_filenames = []
pdf_viewer_commands = ['open', 'evince', 'xdg-open', 'acroread', 'okular', 'atril']

def open_pdf_with_viewer(pdf_path):
    for command in pdf_viewer_commands:
        try:
            subprocess.Popen([command, pdf_path])
            return True
        except FileNotFoundError:
            continue
    return False

def generate_pdf(id_number, preview):
    try:
        # Load the Excel file
        github_file_url = "https://github.com/moestergaard/PreviewRecipe/raw/ebace15b4e687d2ead456565d5a044ca65ea0a65/Data/test.xlsx"

        """
        # Download the Excel file from GitHub
        response = requests.get(github_file_url)
        if response.status_code == 200:
            with tempfile.NamedTemporaryFile(suffix=".xlsx",
                                             delete=False) as temp_xlsx:
                temp_xlsx.write(response.content)
            # Load the downloaded Excel file
            wb = load_workbook(temp_xlsx.name)
            sheet = wb.active
        else:
            print("Failed to download the Excel file from GitHub.")
            exit()
        """

        wb = openpyxl.load_workbook(
            '/Users/martinoestergaard/OneDrive/Martin/Documents/Skole/AU/ITKO/1. semester/DTIV/DTIV Test/TestSenestGitHub.xlsx')
        sheet = wb.active



        # Find the row with the matching ID number
        for row in sheet.iter_rows(values_only=True):
            if row[0] == id_number:
                # Create a PDF with the rest of the row data

                if preview:
                    pdf_filename = BytesIO()
                else:
                    pdf_filename = f'{id_number}.pdf'
                c = canvas.Canvas(pdf_filename, pagesize=letter)
                y = 500  # Starting Y coordinate for content

                image_path = row[len(row) - 1]
                x_image, y_image = 100, 550
                width, height = 200, 200

                if image_path is not None:
                    try:
                        c.drawImage(image_path, x_image, y_image, width,
                                    height)
                    except Exception as e:
                        messagebox.showerror("Error", str(e))

                for value in row[1:-1]:
                    c.drawString(100, y, str(value))
                    y -= 15  # Move up for the next value

                # Closes the current page
                c.showPage()
                # Saves and closes the PDF document in the file
                c.save()

                if preview:
                    with tempfile.NamedTemporaryFile(suffix=".pdf",
                                                     delete=False) as temp_pdf:
                        temp_pdf.write(pdf_filename.getvalue())

                    if not open_pdf_with_viewer(temp_pdf.name):
                        messagebox.showerror("Error",
                                             "Unable to open the PDF. Please install a PDF viewer.")

                else:
                    messagebox.showinfo("PDF Generated",
                                        f"PDF file '{pdf_filename}' has been created.")

                return

        messagebox.showerror("ID not found",
                             f"ID number '{id_number}' not found in the Excel file.")

    except Exception as e:
        messagebox.showerror("Error", str(e))


def preview():
    id_number = id_entry.get()
    if id_number:
        try:
            id_number_int = int(id_number)
        except ValueError:
            messagebox.showerror("Error", "ID number must be an integer.")
            return

        generate_pdf(id_number_int, preview=True)
    else:
        messagebox.showerror("Error", "There must be an ID number.")

def udgiv():
    id_number = id_entry.get()
    if id_number:
        try:
            id_number_int = int(id_number)
        except ValueError:
            messagebox.showerror("Error", "ID number must be an integer.")
            return

        generate_pdf(id_number_int, preview=False)
    else:
        messagebox.showerror("Error", "There must be an ID number.")


# Create the main window
window = tk.Tk()
window.title("Admin Tool Backend")

# Create and place widgets
id_label = tk.Label(window, text="Enter ID Number:")
id_label.pack(pady=10)

id_entry = tk.Entry(window)
id_entry.pack()

preview_button = tk.Button(window, text="Preview", command=preview)
preview_button.pack(pady=5)

udgiv_button = tk.Button(window, text="Udgiv", command=udgiv)
udgiv_button.pack(pady=5)

# Start the tkinter main loop
window.mainloop()
